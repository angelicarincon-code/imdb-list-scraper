# app.py
# Versión genérica y robusta para extraer No., Title, Year, Duration, Age, Rating, Votes
# Copia y pega este archivo completo en tu repo (reemplaza el app.py existente).

import re
import io
import time
from typing import List, Dict, Optional

import requests
from bs4 import BeautifulSoup
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ---------- Config ----------
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9"
}
REQUEST_TIMEOUT = 25
MAX_COL_WIDTH = 50

# ---------- Helpers ----------


def _get_soup(url: str) -> BeautifulSoup:
    r = requests.get(url, headers=HEADERS, timeout=REQUEST_TIMEOUT)
    r.raise_for_status()
    return BeautifulSoup(r.text, "html.parser")


def _to_int(text: Optional[str]) -> Optional[int]:
    if not text:
        return None
    s = re.sub(r"[^\d]", "", str(text))
    return int(s) if s else None


def _to_float(text: Optional[str]) -> Optional[float]:
    if text is None:
        return None
    s = str(text).strip().replace(",", ".")
    m = re.search(r"[-+]?\d*\.?\d+", s)
    try:
        return float(m.group(0)) if m else None
    except Exception:
        return None


def _runtime_to_minutes(rt_text: Optional[str]) -> Optional[int]:
    if not rt_text:
        return None
    rt = str(rt_text)
    # Examples: '2h 22min', '142 min', '142', '1h 30m'
    hours = 0
    minutes = 0
    mh = re.search(r"(\d+)\s*h", rt)
    if mh:
        hours = int(mh.group(1))
    mm = re.search(r"(\d+)\s*min", rt)
    if mm:
        minutes = int(mm.group(1))
    elif not mh:
        # possible single number '142' meaning minutes
        mnum = re.search(r"\b(\d{2,3})\b", rt)
        if mnum:
            return int(mnum.group(1))
    total = hours * 60 + minutes
    return total if total > 0 else None


def _extract_age(text: Optional[str]) -> Optional[str]:
    if not text:
        return None
    t = str(text)
    m = re.search(r"(\d{1,3}\+)", t)
    if m:
        return m.group(1)
    m2 = re.search(r"\b(PG-?\d{1,2}|G|PG|R|NC-17|TV-?MA|TV-?14)\b", t, re.I)
    if m2:
        return m2.group(1).upper()
    return None


def _extract_year(text: Optional[str]) -> Optional[int]:
    if not text:
        return None
    m = re.search(r"(19\d{2}|20\d{2}|21\d{2})", str(text))
    return int(m.group(1)) if m else None


def _clean_votes_text(t: str) -> Optional[int]:
    if not t:
        return None
    # remove non-digits
    s = re.sub(r"[^\d]", "", t)
    return int(s) if s else None


# ---------- Generic parsers ----------


def _gather_candidate_containers(soup: BeautifulSoup) -> List[BeautifulSoup]:
    """
    Devuelve una lista de elementos que parecen contener una entrada (película/serie)
    usando muchos selectores posibles; si no encuentra nada, hace fallback buscando
    anchors a /title/tt... y usando el padre como contenedor.
    """
    selectors = [
        "table.chart tbody tr",                      # Top 250 table rows
        "div.lister-list div.lister-item",           # older list pages
        "div.list_item",                             # other list markup
        "li.ipc-metadata-list-summary-item",         # modern IMDb lists
        "div.ipc-title-card",                        # modern cards
        "div.ranking-list-item",                     # possible variants
        "div.section .list_item",                    # fallback variants
        "div.titleOverview",                         # rare
    ]

    found = []
    for sel in selectors:
        nodes = soup.select(sel)
        if nodes:
            found.extend(nodes)

    # If still empty, try to locate anchors to titles and take their parent block
    if not found:
        anchors = soup.select("a[href^='/title/tt']")
        parents = []
        for a in anchors:
            p = a.find_parent(["li", "div", "tr", "article", "section"]) or a.parent
            if p and p not in parents:
                parents.append(p)
        found = parents

    # Deduplicate preserving order
    seen = set()
    uniq = []
    for el in found:
        sid = str(hash(el))
        if sid not in seen:
            seen.add(sid)
            uniq.append(el)
    return uniq


def _extract_from_container(el: BeautifulSoup) -> Dict:
    """
    Extrae Title, Year, Duration (min), Age, Rating, Votes desde un contenedor HTML.
    Usa múltiples estrategias por campo.
    """
    text = el.get_text(" ", strip=True)

    # ---- Title ----
    title = None
    title_selectors = [
        "td.titleColumn a",
        "a.ipc-title-link, a.ipc-title-link-wrapper",
        "h3.lister-item-header a",
        "h3 a",
        "a[href^='/title/']"
    ]
    for sel in title_selectors:
        t = el.select_one(sel)
        if t and t.get_text(strip=True):
            title = t.get_text(" ", strip=True)
            break
    if title:
        title = re.sub(r"^\d+\.\s*", "", title).strip()

    # ---- Year ----
    year = None
    # common places
    y_candidates = [
        el.select_one("span.secondaryInfo"),
        el.select_one("span.lister-item-year"),
        el.select_one(".year"),
        el.find(string=re.compile(r"\(\d{4}\)"))
    ]
    for yc in y_candidates:
        if yc:
            year = _extract_year(yc.get_text(strip=True) if hasattr(yc, "get_text") else str(yc))
            if year:
                break
    if not year:
        # search in the container text
        year = _extract_year(text)

    # ---- Duration ----
    duration = None
    rt_candidates = [
        el.select_one("span.runtime"),
        el.select_one(".runtime"),
        el.select_one("li.runtime"),
        el.select_one("div.runtime")
    ]
    for rc in rt_candidates:
        if rc and rc.get_text(strip=True):
            duration = _runtime_to_minutes(rc.get_text(strip=True))
            if duration:
                break
    if not duration:
        # search by regex in text
        m = re.search(r"(\d+h\s*\d*min|\d+\s*min|\b\d{2,3}\b\s*min|\b\d{2,3}\b(?!\s*votes))", text, re.I)
        if m:
            duration = _runtime_to_minutes(m.group(1))

    # ---- Age / Certificate ----
    age = None
    # look for explicit certificate spans
    cert = el.find(string=re.compile(r"\b(PG-?\d{1,2}|G|PG|R|NC-17|TV-?MA|TV-?14|\d{1,3}\+)\b", re.I))
    if cert:
        age = _extract_age(cert)
    if not age:
        # try near runtime or in the text
        age = _extract_age(text)

    # ---- Rating ----
    rating = None
    rating_selectors = [
        "td.ratingColumn.imdbRating strong",
        "span.ipc-rating-star--rating",
        "span.aggregate-rating",
        "span[class*='ratingValue']",
        "span[itemprop='ratingValue']",
        "strong"
    ]
    for rs in rating_selectors:
        rnode = el.select_one(rs)
        if rnode and rnode.get_text(strip=True):
            rating = _to_float(rnode.get_text(strip=True))
            if rating is not None:
                break
    # fallback: look for pattern like '9.3/10'
    if rating is None:
        m = re.search(r"(\d\.\d)\s*/\s*10", text)
        if m:
            rating = _to_float(m.group(1))

    # ---- Votes ----
    votes = None
    # common selectors for votes
    nv = el.select_one("span[name='nv']") or el.select_one("span.ipc-rating-star--voteCount") or el.select_one(".votes")
    if nv:
        # nv may be like '1,234' or have data-value attribute
        if nv.has_attr("data-value"):
            votes = _to_int(nv["data-value"])
        else:
            votes = _clean_votes_text(nv.get_text())
    # Some tables store votes in the title attribute of <strong>
    if votes is None:
        strong = el.select_one("td.ratingColumn.imdbRating strong")
        if strong and strong.has_attr("title"):
            # example: '9.2 based on 1,600,000 user ratings'
            m = re.search(r"based on\s+([\d,\.]+)\s+user", strong["title"], re.I)
            if not m:
                m = re.search(r"([\d,\.]+)\s+user", strong["title"], re.I)
            if m:
                votes = _clean_votes_text(m.group(1))
    # fallback: search text for 'votes' numbers
    if votes is None:
        m = re.search(r"([\d,\.]{2,})\s*(?:votes|user ratings|user ratings|ratings)", text, re.I)
        if m:
            votes = _clean_votes_text(m.group(1))
    # final fallback: try to find any large number in text (heuristic)
    if votes is None:
        m = re.search(r"\b(\d{1,3}(?:[.,]\d{3})+)\b", text)
        if m:
            votes = _clean_votes_text(m.group(1))

    return {
        "Title": title or "",
        "Year": year,
        "Duration": duration,   # minutes (int) or None
        "Age": age or "",
        "Rating": rating,
        "Votes": votes
    }


def scrape_imdb_generic(url: str) -> pd.DataFrame:
    """
    Intenta localizar cualquier tipo de listado de IMDb y extraer la info solicitada.
    Funciona con Top 250 (tabla), listas modernas, búsquedas y páginas que tengan anchors a /title/tt...
    """
    soup = _get_soup(url)
    containers = _gather_candidate_containers(soup)

    rows = []
    for el in containers:
        try:
            data = _extract_from_container(el)
            # skip empty titles
            if data.get("Title"):
                rows.append(data)
        except Exception:
            continue
        # cortesía para reducir bloqueos en requests si se recorren muchos items en remoto
        # (aquí no hacemos requests por item, solo parse local, así que no dormimos)

    # Si no encontró contenedores/filas, intentar parse genérico por anchors a /title/tt
    if not rows:
        anchors = soup.select("a[href^='/title/tt']")
        seen = set()
        for idx, a in enumerate(anchors, start=1):
            parent = a.find_parent(["li", "div", "tr", "article", "section"]) or a.parent
            sid = str(hash(parent))
            if sid in seen:
                continue
            seen.add(sid)
            try:
                data = _extract_from_container(parent)
                if data.get("Title"):
                    rows.append(data)
            except Exception:
                continue

    # Normalizar columnas y DataFrame
    cols = ["No.", "Title", "Year", "Duration", "Age", "Rating", "Votes"]
    if not rows:
        df = pd.DataFrame(columns=cols)
    else:
        df = pd.DataFrame(rows)
        # ensure columns exist
        for c in ["Title", "Year", "Duration", "Age", "Rating", "Votes"]:
            if c not in df.columns:
                df[c] = None
        # keep original order when possible; remove exact duplicates by Title+Year
        df["Title_norm"] = df["Title"].str.strip().str.lower()
        df["Year_norm"] = df["Year"].fillna("").astype(str)
        df = df.drop_duplicates(subset=["Title_norm", "Year_norm"], keep="first")
        df = df.drop(columns=["Title_norm", "Year_norm"])
        df.reset_index(drop=True, inplace=True)
        # add No.
        df.insert(0, "No.", range(1, len(df) + 1))

    return df


# ---------- Excel export ( estilizado simple ) ----------


def df_to_excel_bytes(df: pd.DataFrame, sheet_name: str = "IMDb List") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    headers = list(df.columns)
    header_font = Font(bold=True, color="000000")
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_align = Alignment(vertical="top", wrap_text=True)
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[1].height = 20

    # Rows
    for _, row in df.iterrows():
        values = []
        for c in headers:
            v = row[c]
            # if Duration is integer, write as int
            if c == "Duration" and pd.notna(v):
                try:
                    v = int(v)
                except Exception:
                    pass
            values.append(v)
        ws.append(values)

    # Cell formatting & column widths
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = body_align
            cell.border = border

    for col_idx, col_name in enumerate(headers, start=1):
        max_len = len(col_name)
        for cell in ws[get_column_letter(col_idx)]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, MAX_COL_WIDTH)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()


# ---------- Streamlit UI ----------

st.set_page_config(page_title="IMDb Generic Scraper", page_icon="🎬", layout="wide")
st.title("🎬 IMDb — Generic List Scraper")
st.caption("Pega cualquier URL de IMDb (Top, listas modernas, búsquedas). Intento extraer Title, Year, Duration, Age, Rating y Votes de forma genérica.")

url = st.text_input("Pega la URL de IMDb aquí (ej. https://www.imdb.com/chart/top/)", value="")

col1, col2 = st.columns([1, 3])
with col1:
    start = st.button("Extraer datos")
with col2:
    info = st.empty()

if start:
    if not url or not url.strip():
        st.error("Por favor pega una URL de IMDb válida.")
        st.stop()

    info.info("Descargando y analizando la página...")
    t0 = time.time()
    try:
        df = scrape_imdb_generic(url.strip())
    except Exception as e:
        st.error(f"Error al procesar la URL: {e}")
        st.stop()
    elapsed = time.time() - t0
    info.success(f"Listo en {elapsed:.2f}s — {len(df):,} filas encontradas")

    if df.empty:
        st.warning("No se encontraron entradas reconocibles en esta URL. Intenta con la versión en inglés (ej. https://www.imdb.com/chart/top/) o comparte la URL aquí para que la revise.")
    else:
        # Mostrar y ofrecer descarga
        st.subheader("Resultados")
        # Presentar Duration en forma legible (minutos)
        display_df = df.copy()
        display_df["Duration"] = display_df["Duration"].apply(lambda x: f\"{int(x)} min\" if pd.notna(x) and str(x).isdigit() else (f\"{int(x)} min\" if pd.notna(x) and isinstance(x, (int, float)) else (x if pd.notna(x) else \"\")))
        st.dataframe(display_df, use_container_width=True, hide_index=True)

        xlsx_bytes = df_to_excel_bytes(df)
        st.download_button(
            label="⬇️ Descargar Excel",
            data=xlsx_bytes,
            file_name="imdb_list.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

st.markdown("---")
st.caption("Si una URL no funciona, pégala aquí y la reviso. Este extractor intenta muchas reglas para ser lo más genérico posible.")
